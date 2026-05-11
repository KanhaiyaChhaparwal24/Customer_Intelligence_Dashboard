import sys
from pathlib import Path

files = [
    Path('Shopify_Order_Data.xlsx'),
    Path('Warranty_Registration_Data.xlsx')
]

try:
    import openpyxl
except Exception as e:
    print('openpyxl not available:', e)
    sys.exit(2)

for f in files:
    print('\n' + '='*60)
    print('FILE:', f)
    if not f.exists():
        print('  MISSING')
        continue
    wb = openpyxl.load_workbook(str(f), read_only=True, data_only=True)
    print('  Sheets:', wb.sheetnames)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print('\n  Sheet:', sheet)
        rows = ws.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            print('    EMPTY SHEET')
            continue
        print('    Headers:', headers)
        print('    Sample rows:')
        for i, r in enumerate(rows, start=1):
            print('     ', r)
            if i >= 10:
                break
    wb.close()

print('\nDone')
