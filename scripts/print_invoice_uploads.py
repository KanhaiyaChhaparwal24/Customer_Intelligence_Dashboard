from openpyxl import load_workbook

wb = load_workbook('sheets/Warranty_Registration_Data.xlsx', read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
headers = next(ws.iter_rows(values_only=True))
idx = headers.index('Invoice Upload')
for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
    if i == 1:
        continue
    print(f'{i-1}: {row[idx]}')
wb.close()
