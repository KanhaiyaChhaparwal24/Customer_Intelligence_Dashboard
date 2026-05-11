from openpyxl import load_workbook

wb = load_workbook('sheets/Warranty_Registration_Data.xlsx', read_only=False, data_only=True)
ws = wb[wb.sheetnames[0]]
headers = [c.value for c in ws[1]]
idx = headers.index('Invoice Upload') + 1
for r in range(2, min(ws.max_row, 12) + 1):
    cell = ws.cell(r, idx)
    print(r - 1, 'value=', cell.value, 'hyperlink=', getattr(cell.hyperlink, 'target', None))
wb.close()
