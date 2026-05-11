"""
sheets_service.py
Incremental Google Sheets reader with fuzzy column mapping.
Scales to 50k+ rows by reading only rows beyond last_processed_row.
"""
import logging
from typing import List, Dict, Tuple, Optional
import os
from pathlib import Path

import gspread
from auth import get_credentials
from config import GOOGLE_SHEET_NAME, COLUMN_MAPPING, USE_LOCAL_SHEETS, LOCAL_SHEETS_DIR

try:
    import openpyxl
except Exception:
    openpyxl = None

logger = logging.getLogger(__name__)


def _fuzzy_match_column(headers: List[str], field: str) -> Optional[str]:
    """Find the best matching header for a logical field name."""
    candidates = COLUMN_MAPPING.get(field, [field])
    # 1. Exact case-insensitive match
    for candidate in candidates:
        for header in headers:
            if header.strip().lower() == candidate.strip().lower():
                return header
    # 2. Substring / partial match fallback
    for candidate in candidates:
        for header in headers:
            if candidate.lower() in header.lower() or header.lower() in candidate.lower():
                return header
    return None


def _build_col_map(headers: List[str], fields: List[str]) -> Dict[str, str]:
    """Build {field_name: actual_header} mapping for a list of fields."""
    col_map = {}
    for field in fields:
        matched = _fuzzy_match_column(headers, field)
        if matched:
            col_map[field] = matched
        else:
            logger.debug(f"Column not found for field '{field}' in headers: {headers}")
    return col_map


def _open_sheet():
    creds = get_credentials()
    gc = gspread.authorize(creds)
    return gc.open(GOOGLE_SHEET_NAME)


# ─────────────────────────────────────────────────────────────────────────────
# Warranty Sheet (Tab 0 / local: Warranty_Registration_Data.xlsx)
# Customer warranty registration form with invoice link to Drive or local file
# Exact headers: Timestamp  Email  Phone  Brand  Product Name  Colour  Size  Invoice Upload
# ─────────────────────────────────────────────────────────────────────────────

WARRANTY_FIELDS = [
    "timestamp", "email", "phone", "brand",
    "product_name", "colour", "size", "invoice_link",
]


def read_warranty_rows(last_row: int = 0) -> Tuple[List[Dict], int]:
    """
    Read warranty sheet incrementally.
    Returns (list_of_normalized_rows, total_row_count).
    Only rows with row_number > last_row are returned.
    Row numbers are 1-based (row 1 = first data row after header).
    """
    # Prefer local XLSX when configured and available
    if USE_LOCAL_SHEETS and openpyxl is not None:
        wb_path = Path(LOCAL_SHEETS_DIR) / "Warranty_Registration_Data.xlsx"
        if wb_path.exists():
            wb = openpyxl.load_workbook(str(wb_path), read_only=False, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows_iter = ws.iter_rows()
            try:
                header_cells = next(rows_iter)
            except StopIteration:
                wb.close()
                return [], 0

            headers = [str(c.value or "") for c in header_cells]
            col_map = _build_col_map(headers, WARRANTY_FIELDS)

            rows: List[Dict] = []
            total_data_rows = 0
            for row_idx, row_cells in enumerate(rows_iter, start=1):
                total_data_rows += 1
                if row_idx <= last_row:
                    continue
                row_cells = list(row_cells) + [None] * (len(headers) - len(row_cells))
                row_dict = {headers[i]: row_cells[i] for i in range(len(headers))}

                normalized: Dict = {"_row_number": row_idx}
                for field, col_name in col_map.items():
                    cell = row_dict.get(col_name)
                    if cell is None:
                        normalized[field] = ""
                        continue
                    if field == "invoice_link":
                        link = getattr(cell.hyperlink, "target", None)
                        normalized[field] = str(link or cell.value or "").strip()
                    else:
                        val = cell.value
                        if isinstance(val, float) and val.is_integer():
                            val = int(val)
                        normalized[field] = str(val if val is not None else "").strip()

                has_data = any(v for k, v in normalized.items() if k != "_row_number")
                if not has_data:
                    continue

                rows.append(normalized)

            wb.close()
            logger.info(
                f"Local warranty sheet: {total_data_rows} total rows, {len(rows)} new rows after row {last_row}"
            )
            return rows, total_data_rows

    # Fallback to Google Sheets
    try:
        sheet = _open_sheet()
        ws = sheet.get_worksheet(0)
        all_values = ws.get_all_values()

        if not all_values or len(all_values) < 2:
            return [], 0

        headers = all_values[0]
        col_map = _build_col_map(headers, WARRANTY_FIELDS)
        total_data_rows = len(all_values) - 1

        rows: List[Dict] = []
        for row_idx, row_values in enumerate(all_values[1:], start=1):
            if row_idx <= last_row:
                continue

            # Pad short rows
            row_values = row_values + [""] * (len(headers) - len(row_values))
            row_dict = dict(zip(headers, row_values))

            normalized: Dict = {"_row_number": row_idx}
            for field, col_name in col_map.items():
                normalized[field] = row_dict.get(col_name, "").strip()

            # Skip entirely empty rows
            has_data = any(
                v for k, v in normalized.items() if k != "_row_number"
            )
            if not has_data:
                continue

            rows.append(normalized)

        logger.info(
            f"Warranty sheet: {total_data_rows} total rows, "
            f"{len(rows)} new rows after row {last_row}"
        )
        return rows, total_data_rows

    except Exception as e:
        logger.error(f"Error reading warranty sheet: {e}")
        raise


# ─────────────────────────────────────────────────────────────────────────────
# Shopify Sheet (Tab 1 / local: Shopify_Order_Data.xlsx)
# D2C orders from Shopify platform export (79-column CSV format)
# Column priority order (first non-empty wins per field):
#   order_id       -> "Id" (Shopify numeric ID) else email+date composite
#   customer_name  -> "Billing Name" > "Name"
#   email          -> "Email"
#   phone          -> "Billing Phone" > "Shipping Phone" > "Phone"
#   city           -> "Billing City" > "Shipping City"
#   state          -> "Billing Province Name" > "Billing Province" > "Shipping Province Name"
#   total          -> "Total" > "Subtotal"
#   product        -> "Lineitem name"
#   created_at     -> "Paid at" > "Created at"
#   payment_method -> "Payment Method"
# ─────────────────────────────────────────────────────────────────────────────

def read_shopify_rows() -> List[Dict]:
    """
    Read all Shopify orders from Tab 2.
    Uses direct case-insensitive column-name lookup with priority fallbacks
    to correctly handle the full 79-column Shopify CSV export format.
    """
    # Local XLSX path preferred
    if USE_LOCAL_SHEETS and openpyxl is not None:
        wb_path = Path(LOCAL_SHEETS_DIR) / "Shopify_Order_Data.xlsx"
        if wb_path.exists():
            wb = openpyxl.load_workbook(str(wb_path), read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                headers = next(rows_iter)
            except StopIteration:
                wb.close()
                return []

            headers = [h if h is not None else "" for h in headers]
            hmap: Dict[str, int] = {h.strip().lower(): i for i, h in enumerate(headers)}

            def get(row_vals: List[str], *col_names: str) -> str:
                for name in col_names:
                    idx = hmap.get(name.lower())
                    if idx is not None and idx < len(row_vals):
                        raw = row_vals[idx]
                        val = str(raw).strip() if raw is not None else ""
                        if val:
                            return val
                return ""

            rows: List[Dict] = []
            for row_values in rows_iter:
                row_values = list(row_values) + [""] * max(0, len(headers) - len(row_values))

                email      = get(row_values, "email")
                cname      = get(row_values, "billing name", "name")
                shop_id    = get(row_values, "id")
                phone      = get(row_values, "billing phone", "shipping phone", "phone")
                city       = get(row_values, "billing city", "shipping city")
                state      = get(row_values, "billing province name", "billing province",
                                            "shipping province name", "shipping province")
                total      = get(row_values, "total", "subtotal")
                product    = get(row_values, "lineitem name")
                created_at = get(row_values, "paid at", "created at")
                payment    = get(row_values, "payment method")

                if not email and not cname:
                    continue

                if shop_id:
                    stable_id = shop_id
                else:
                    identity = email or (cname or "").lower()
                    stable_id = f"{identity}_{created_at}" if created_at else identity

                rows.append({
                    "order_id":       stable_id,
                    "customer_name":  cname or email,
                    "email":          email.lower() if email else "",
                    "phone":          phone,
                    "city":           city,
                    "state":          state,
                    "total":          total,
                    "product":        product,
                    "created_at":     created_at,
                    "payment_method": payment,
                })

            wb.close()
            logger.info(f"Local Shopify sheet: {len(rows)} orders read from {wb_path}")
            return rows

    # Fallback to Google Sheets
    try:
        sheet = _open_sheet()
        ws = sheet.get_worksheet(1)
        all_values = ws.get_all_values()

        if not all_values or len(all_values) < 2:
            return []

        headers = all_values[0]
        # Build case-insensitive header -> index map
        hmap: Dict[str, int] = {h.strip().lower(): i for i, h in enumerate(headers)}

        def get(row_vals: List[str], *col_names: str) -> str:
            """Try column names in priority order, return first non-empty value."""
            for name in col_names:
                idx = hmap.get(name.lower())
                if idx is not None and idx < len(row_vals):
                    raw = row_vals[idx]
                    val = str(raw).strip() if raw is not None else ""
                    if val:
                        return val
            return ""

        rows: List[Dict] = []
        for row_values in all_values[1:]:
            row_values = row_values + [""] * max(0, len(headers) - len(row_values))

            email      = get(row_values, "email")
            cname      = get(row_values, "billing name", "name")
            shop_id    = get(row_values, "id")
            phone      = get(row_values, "billing phone", "shipping phone", "phone")
            city       = get(row_values, "billing city", "shipping city")
            state      = get(row_values, "billing province name", "billing province",
                                        "shipping province name", "shipping province")
            total      = get(row_values, "total", "subtotal")
            product    = get(row_values, "lineitem name")
            created_at = get(row_values, "paid at", "created at")
            payment    = get(row_values, "payment method")

            # Skip rows with no useful identity
            if not email and not cname:
                continue

            # Build a stable unique order_id:
            # Prefer Shopify's numeric Id; fallback to email+date composite key
            if shop_id:
                stable_id = shop_id
            else:
                identity = email or cname.lower()
                stable_id = f"{identity}_{created_at}" if created_at else identity

            rows.append({
                "order_id":       stable_id,
                "customer_name":  cname or email,
                "email":          email.lower() if email else "",
                "phone":          phone,
                "city":           city,
                "state":          state,
                "total":          total,
                "product":        product,
                "created_at":     created_at,
                "payment_method": payment,
            })

        logger.info(f"Shopify sheet: {len(rows)} orders read from {len(all_values)-1} raw rows")
        return rows

    except Exception as e:
        logger.error(f"Error reading Shopify sheet: {e}")
        raise
